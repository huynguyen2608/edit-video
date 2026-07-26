"""Kho trạng thái bằng FILE EXCEL (không dùng database).

Toàn bộ dữ liệu để ở local trong MỘT workbook .xlsx, chia theo sheet:
  - sheet "videos":   mỗi video một dòng (trạng thái tải/edit, đường dẫn…)
  - sheet "channels": mỗi kênh một dòng (đã quét lần cuối…)
Mở bằng Excel để xem/quản lý trực tiếp.

Thiết kế thay cho SQLite nhưng GIỮ NGUYÊN interface (ExcelStore có đủ method như
Database cũ) nên scheduler/pipeline/GUI/worker không phải sửa logic.

Đồng bộ đa luồng: đây là app 1 tiến trình (scheduler nền + QThread quét/edit) nên
một RLock trong tiến trình là đủ. Mỗi thao tác ghi: sửa model trong RAM rồi lưu
nguyên workbook (ghi ra file tạm rồi os.replace -> không hỏng file nếu crash giữa
chừng). Với 15-30 video/ngày tải trọng rất nhẹ.

Hai trục trạng thái độc lập (giữ như cũ):
  download_status: pending -> downloading -> downloaded | failed
  edit_status:     pending -> processing  -> done       | failed
"""
from __future__ import annotations

import os
import threading
import time
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Optional

from openpyxl import Workbook, load_workbook

from .logging_setup import get_logger

log = get_logger("store")

VIDEO_COLS = [
    "video_id", "channel_id", "channel_name", "title", "url", "published",
    "download_status", "download_path", "edit_status", "edited_at", "error", "added_at",
    # ---- Queue Manager (theo dõi tiến trình edit) ----
    "job_status", "stage", "progress", "duration", "resolution", "fps",
]
CHANNEL_COLS = ["channel_id", "name", "url", "last_scanned"]
# Log XUẤT BẢN: mỗi lần export ghi 1 dòng (3 file đầu ra + thời điểm).
EXPORT_COLS = ["video_id", "channel_name", "full_path", "short_path",
               "content_txt", "srt_path", "output_dir", "exported_at"]
# Log LỊCH SỬ: dòng thời gian các sự kiện quan trọng (tải/edit/lỗi…).
EVENT_COLS = ["time", "level", "source", "message"]

VIDEOS_SHEET = "videos"
CHANNELS_SHEET = "channels"
EXPORTS_SHEET = "exports"
EVENTS_SHEET = "events"

_EVENTS_MAX = 10000  # giữ tối đa để file không phình vô hạn


@dataclass
class VideoRow:
    video_id: str
    channel_id: str
    channel_name: str
    title: str
    url: str
    published: str
    download_status: str
    download_path: Optional[str]
    edit_status: str


def _now() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


class ExcelStore:
    def __init__(self, path: str | Path):
        self.path = str(path)
        Path(self.path).parent.mkdir(parents=True, exist_ok=True)
        self._lock = threading.RLock()
        self._videos: dict[str, dict] = {}    # video_id -> hàng
        self._channels: dict[str, dict] = {}  # channel_id -> hàng
        self._exports: list[dict] = []        # log xuất bản (append-only)
        self._events: list[dict] = []         # log lịch sử (append-only, cắt bớt cũ)
        self._load()
        if not Path(self.path).exists():
            self._save()  # tạo file rỗng có sẵn header để mở bằng Excel được ngay

    # ---------------- I/O ----------------
    def _load(self) -> None:
        p = Path(self.path)
        if not p.exists():
            return
        wb = load_workbook(self.path, read_only=True, data_only=True)
        if VIDEOS_SHEET in wb.sheetnames:
            self._videos = _read_sheet(wb[VIDEOS_SHEET], VIDEO_COLS, "video_id")
        if CHANNELS_SHEET in wb.sheetnames:
            self._channels = _read_sheet(wb[CHANNELS_SHEET], CHANNEL_COLS, "channel_id")
        if EXPORTS_SHEET in wb.sheetnames:
            self._exports = _read_sheet_list(wb[EXPORTS_SHEET], EXPORT_COLS)
        if EVENTS_SHEET in wb.sheetnames:
            self._events = _read_sheet_list(wb[EVENTS_SHEET], EVENT_COLS)
        wb.close()

    def _save(self) -> None:
        wb = Workbook()
        vs = wb.active
        vs.title = VIDEOS_SHEET
        vs.append(VIDEO_COLS)
        for row in self._videos.values():
            vs.append([row.get(c) for c in VIDEO_COLS])
        cs = wb.create_sheet(CHANNELS_SHEET)
        cs.append(CHANNEL_COLS)
        for row in self._channels.values():
            cs.append([row.get(c) for c in CHANNEL_COLS])
        es = wb.create_sheet(EXPORTS_SHEET)
        es.append(EXPORT_COLS)
        for row in self._exports:
            es.append([row.get(c) for c in EXPORT_COLS])
        evs = wb.create_sheet(EVENTS_SHEET)
        evs.append(EVENT_COLS)
        for row in self._events:
            evs.append([row.get(c) for c in EVENT_COLS])
        # ghi file tạm rồi thay thế nguyên tử -> không hỏng workbook nếu crash giữa chừng
        tmp = f"{self.path}.tmp"
        wb.save(tmp)
        wb.close()
        # Nếu data.xlsx đang mở trong Excel -> os.replace bị khóa. Thử lại vài lần;
        # vẫn không được thì bỏ qua (trạng thái vẫn còn trong RAM, sẽ ghi ở lần sau)
        # để không làm hỏng luồng tải/edit đang chạy.
        last: Exception | None = None
        for _ in range(4):
            try:
                os.replace(tmp, self.path)
                return
            except PermissionError as e:
                last = e
                time.sleep(0.15)
        log.warning("Chưa ghi được %s (đang mở ở ứng dụng khác?) — sẽ thử lại lần lưu sau. %s",
                    self.path, last)
        try:
            os.remove(tmp)
        except OSError:
            pass

    # ---------------- channels ----------------
    def upsert_channel(self, channel_id: str, name: str, url: str = "") -> None:
        with self._lock:
            row = self._channels.get(channel_id, {"channel_id": channel_id})
            row["name"] = name
            row["url"] = url
            self._channels[channel_id] = row
            self._save()

    def mark_channel_scanned(self, channel_id: str) -> None:
        with self._lock:
            row = self._channels.setdefault(channel_id, {"channel_id": channel_id})
            row["last_scanned"] = _now()
            self._save()

    # ---------------- videos ----------------
    def video_exists(self, video_id: str) -> bool:
        with self._lock:
            return video_id in self._videos

    def add_discovered(self, video_id: str, channel_id: str, channel_name: str,
                       title: str, url: str, published: str) -> bool:
        """Thêm video mới phát hiện. Trả True nếu là video MỚI (chưa từng có)."""
        with self._lock:
            if video_id in self._videos:
                return False
            self._videos[video_id] = {
                "video_id": video_id, "channel_id": channel_id,
                "channel_name": channel_name, "title": title, "url": url,
                "published": published, "download_status": "pending",
                "download_path": None, "edit_status": "pending",
                "edited_at": None, "error": None, "added_at": _now(),
            }
            self._save()
            return True

    def add_local_video(self, video_id: str, channel_name: str, title: str,
                        path: str) -> bool:
        """Nạp video từ FOLDER local (coi như đã 'downloaded'), cho luồng edit độc lập.

        Trả True nếu MỚI thêm. Nếu đã có: chỉ cập nhật đường dẫn (phòng khi di chuyển
        file) và GIỮ nguyên edit_status (không edit lại video đã 'done').
        """
        with self._lock:
            row = self._videos.get(video_id)
            if row:
                row["download_path"] = path
                self._save()
                return False
            self._videos[video_id] = {
                "video_id": video_id, "channel_id": "local",
                "channel_name": channel_name, "title": title, "url": path,
                "published": _now(), "download_status": "downloaded",
                "download_path": path, "edit_status": "pending",
                "edited_at": None, "error": None, "added_at": _now(),
            }
            self._save()
            return True

    def claim_download(self, video_id: str) -> bool:
        """Chuyển pending -> downloading NGUYÊN TỬ. True nếu giành được quyền tải.

        RLock đảm bảo hai phiên quét song song không cùng tải một video: chỉ phiên
        đầu thấy trạng thái 'pending' và đổi được; phiên sau thấy 'downloading' -> bỏ.
        """
        with self._lock:
            row = self._videos.get(video_id)
            if not row or row.get("download_status") != "pending":
                return False
            row["download_status"] = "downloading"
            self._save()
            return True

    def set_download_status(self, video_id: str, status: str,
                            path: str | None = None, error: str | None = None) -> None:
        with self._lock:
            row = self._videos.get(video_id)
            if not row:
                return
            row["download_status"] = status
            if path is not None:            # COALESCE: chỉ ghi đè khi có giá trị mới
                row["download_path"] = path
            row["error"] = error
            self._save()

    def transition_download_status(self, video_id: str, expected, status: str) -> bool:
        """Đổi trạng thái có điều kiện để tránh ghi đè kết quả từ worker vừa hoàn tất."""
        allowed = {expected} if isinstance(expected, str) else set(expected)
        with self._lock:
            row = self._videos.get(video_id)
            if not row or row.get("download_status") not in allowed:
                return False
            row["download_status"] = status
            self._save()
            return True

    def set_edit_status(self, video_id: str, status: str, error: str | None = None) -> None:
        with self._lock:
            row = self._videos.get(video_id)
            if not row:
                return
            row["edit_status"] = status
            if status == "done":
                row["edited_at"] = _now()
            row["error"] = error
            self._save()

    def claim_edit(self, video_id: str) -> bool:
        """Nhận quyền edit nguyên tử: downloaded/pending -> processing."""
        with self._lock:
            row = self._videos.get(video_id)
            if (not row or row.get("download_status") != "downloaded" or
                    row.get("edit_status") != "pending"):
                return False
            row["edit_status"] = "processing"
            row["error"] = None
            self._save()
            return True

    def pending_downloads(self) -> list[VideoRow]:
        return self._query(lambda r: r.get("download_status") == "pending",
                           key="published", reverse=True)

    def pending_edits(self) -> list[VideoRow]:
        """Video đã tải xong và chưa edit — đầu vào của Module 2."""
        return self._query(
            lambda r: r.get("download_status") == "downloaded"
            and r.get("edit_status") == "pending",
            key="published", reverse=True,
        )

    def all_videos(self) -> list[VideoRow]:
        return self._query(lambda r: True, key="added_at", reverse=True)

    def all_video_rows(self) -> list[dict]:
        """Toàn bộ video dạng dict ĐẦY ĐỦ cột (gồm job_status/stage/progress…) cho Queue."""
        with self._lock:
            rows = [dict(r) for r in self._videos.values()]
        rows.sort(key=lambda r: (r.get("added_at") or ""), reverse=True)
        return rows

    def get_video(self, video_id: str) -> Optional[VideoRow]:
        with self._lock:
            r = self._videos.get(video_id)
            if not r:
                return None
            return VideoRow(**{c: r.get(c) for c in
                               ("video_id", "channel_id", "channel_name", "title", "url",
                                "published", "download_status", "download_path", "edit_status")})

    # ---------------- log lịch sử & xuất bản ----------------
    def log_event(self, source: str, message: str, level: str = "INFO") -> None:
        """Ghi 1 sự kiện vào sheet 'events' (lịch sử tải/edit/lỗi…)."""
        with self._lock:
            self._events.append({"time": _now(), "level": level,
                                 "source": source, "message": str(message)})
            if len(self._events) > _EVENTS_MAX:
                self._events = self._events[-_EVENTS_MAX:]
            self._save()

    def log_export(self, video_id: str, channel_name: str, output_dir: str,
                   full_path: str | None = None, short_path: str | None = None,
                   content_txt: str | None = None, srt_path: str | None = None) -> None:
        """Ghi 1 dòng vào sheet 'exports' + 1 sự kiện lịch sử."""
        with self._lock:
            self._exports.append({
                "video_id": video_id, "channel_name": channel_name,
                "full_path": full_path, "short_path": short_path,
                "content_txt": content_txt, "srt_path": srt_path,
                "output_dir": output_dir, "exported_at": _now(),
            })
            self._events.append({"time": _now(), "level": "INFO", "source": "export",
                                 "message": f"export {video_id} -> {output_dir}"})
            if len(self._events) > _EVENTS_MAX:
                self._events = self._events[-_EVENTS_MAX:]
            self._save()

    def recent_exports(self) -> list[dict]:
        with self._lock:
            return list(self._exports)

    def recent_events(self) -> list[dict]:
        with self._lock:
            return list(self._events)

    def clear_history(self) -> None:
        """Xoá toàn bộ log lịch sử + xuất bản (sheet 'events' & 'exports').

        KHÔNG đụng tới sheet 'videos'/'channels' hay file video đã tải/xuất.
        """
        with self._lock:
            self._exports.clear()
            self._events.clear()
            self._save()

    def _query(self, pred, key: str, reverse: bool) -> list[VideoRow]:
        with self._lock:
            rows = [r for r in self._videos.values() if pred(r)]
        rows.sort(key=lambda r: (r.get(key) or ""), reverse=reverse)
        return [VideoRow(**{c: r.get(c) for c in
                            ("video_id", "channel_id", "channel_name", "title", "url",
                             "published", "download_status", "download_path", "edit_status")})
                for r in rows]

    def reset_failed_downloads(self) -> int:
        """Đưa các video download_status='failed' về 'pending' để thử tải lại. Trả số lượng."""
        with self._lock:
            n = 0
            for r in self._videos.values():
                if r.get("download_status") == "failed":
                    r["download_status"] = "pending"
                    r["error"] = None
                    n += 1
            if n:
                self._save()
            return n

    def reset_for_redownload(self, video_id: str) -> bool:
        """Đưa 1 video về tải lại TỪ ĐẦU: xóa path cũ, pending cả tải lẫn edit.

        Dùng cho 'Tải lại video lỗi' và đối chiếu file mất. Trả True nếu có video.
        """
        with self._lock:
            row = self._videos.get(video_id)
            if not row:
                return False
            row["download_status"] = "pending"
            row["download_path"] = ""
            row["edit_status"] = "pending"
            row["error"] = None
            row["edited_at"] = None
            # Dọn trạng thái hàng đợi cũ (Removed/Failed/Paused…) để video quay lại
            # 'Đang chờ' -> hiện trong Hàng đợi và được tự biên tập sau khi tải lại.
            row["job_status"] = ""
            row["stage"] = ""
            row["progress"] = 0
            self._save()
            return True

    def resume_paused_downloads(self, video_ids=None) -> int:
        """Đưa video tải tạm dừng về pending; file .part được yt-dlp dùng để tải tiếp."""
        wanted = set(video_ids) if video_ids is not None else None
        with self._lock:
            n = 0
            for video_id, row in self._videos.items():
                if row.get("download_status") == "paused" and (wanted is None or video_id in wanted):
                    row["download_status"] = "pending"
                    row["error"] = None
                    n += 1
            if n:
                self._save()
            return n

    def reset_stuck(self) -> None:
        """Gọi lúc khởi động: video kẹt ở trạng thái *ing (do crash) -> cho chạy lại."""
        with self._lock:
            changed = False
            for r in self._videos.values():
                if r.get("download_status") == "downloading":
                    r["download_status"] = "paused"; changed = True
                if r.get("edit_status") == "processing":
                    r["edit_status"] = "pending"; changed = True
            if changed:
                self._save()

    # ---------------- Queue Manager ----------------
    def remove_videos(self, video_ids) -> int:
        """Xoá các video khỏi kho (không xoá file output). Trả số dòng đã xoá."""
        with self._lock:
            n = 0
            for vid in list(video_ids):
                if self._videos.pop(vid, None) is not None:
                    n += 1
            if n:
                self._save()
            return n

    def remove_incomplete_remote_videos(self) -> int:
        """Dọn bản ghi tải từ xa chưa hoàn tất; giữ local và mọi video đã tải.

        Chỉ xóa bản ghi trong kho, không xóa file video hay kết quả đầu ra.
        """
        with self._lock:
            targets = [
                video_id for video_id, row in self._videos.items()
                if not str(video_id).startswith("local_")
                and row.get("download_status") != "downloaded"
            ]
            for video_id in targets:
                self._videos.pop(video_id, None)
            if targets:
                self._save()
            return len(targets)

    def update_job(self, video_id: str, *, persist: bool = True, **fields) -> None:
        """Cập nhật field queue (job_status/stage/progress/duration/resolution/fps…).

        persist=False: chỉ đổi trong RAM (dùng cho progress-tick liên tục, tránh ghi
        file quá nhiều). persist=True: ghi file (dùng ở mốc đổi trạng thái).
        """
        with self._lock:
            row = self._videos.get(video_id)
            if not row:
                return
            for k, v in fields.items():
                if v is not None:
                    row[k] = v
            if persist:
                self._save()

    def resume_queue(self) -> None:
        """Lúc mở lại app: job đang chạy (Preparing/Processing/Rendering) -> Interrupted
        để tiếp tục; Completed giữ nguyên (không chạy lại). Đồng thời reset edit_status
        'processing' -> 'pending'."""
        with self._lock:
            for r in self._videos.values():
                if r.get("job_status") in ("Preparing", "Processing", "Rendering"):
                    r["job_status"] = "Interrupted"
                if r.get("edit_status") == "processing":
                    r["edit_status"] = "pending"
            self._save()


def _read_sheet(ws, cols: list[str], key: str) -> dict[str, dict]:  # noqa: E302
    """Đọc sheet -> dict theo cột header thực tế (bỏ qua cột lạ, thiếu -> None)."""
    it = ws.iter_rows(values_only=True)
    try:
        header = list(next(it))
    except StopIteration:
        return {}
    idx = {name: i for i, name in enumerate(header) if name in cols}
    out: dict[str, dict] = {}
    for raw in it:
        if raw is None:
            continue
        row = {c: (raw[idx[c]] if c in idx and idx[c] < len(raw) else None) for c in cols}
        kv = row.get(key)
        if kv in (None, ""):
            continue
        row[key] = str(kv)
        out[str(kv)] = row
    return out


def _read_sheet_list(ws, cols: list[str]) -> list[dict]:
    """Đọc sheet append-only (log) -> list dict theo header thực tế."""
    it = ws.iter_rows(values_only=True)
    try:
        header = list(next(it))
    except StopIteration:
        return []
    idx = {name: i for i, name in enumerate(header) if name in cols}
    out: list[dict] = []
    for raw in it:
        if raw is None or all(v in (None, "") for v in raw):
            continue
        out.append({c: (raw[idx[c]] if c in idx and idx[c] < len(raw) else None) for c in cols})
    return out
